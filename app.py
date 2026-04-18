import io
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FY_START = pd.Timestamp('2026-04-01')
FY_END = pd.Timestamp('2027-03-31')
MONTH_ENDS = list(pd.date_range(FY_START, FY_END, freq='ME'))
MONTH_LABELS = [d.strftime('%b-%y') for d in MONTH_ENDS]

BASE_REQUIRED_COLS = {
    'Assets ID': ['assets id', 'asset id', 'assetid'],
    'Assets Description': ['assets discreption', 'asset description', 'assets description'],
    'Assets Class': ['assets class', 'asset class'],
    'Purchase Date': ['purchase date', 'date of purchase'],
    'Gross Block Closing Value': ['gross block closing value', 'gross block', 'cost', 'asset value'],
    'Salvage Value': ['salvage value', 'scrap value'],
    'Depreciation Method': ['depreciation method', 'method'],
    'Assets Location': ['assets location', 'location'],
    'Depreciation Rate': ['depreciation rate', 'rate'],
    'Assets Life': ['assets life', 'useful life', 'life'],
    'Closing Gross Block': ['closing gross block'],
    'Closing  Accummlated Depreciation': ['closing accummlated depreciation', 'closing accumulated depreciation'],
}

ADDITION_TEMPLATE_COLS = [
    'Addition Date', 'Assets ID', 'Assets Description', 'Assets Class', 'Assets Location',
    'Gross Block Closing Value', 'Salvage Value', 'Depreciation Method', 'Depreciation Rate',
    'Assets Life', 'Cap At Scrap Value', 'Vendor', 'Invoice Number', 'Remark'
]

DISPOSAL_TEMPLATE_COLS = ['Disposal Date', 'Assets ID', 'Disposed Amount', 'Sale Value', 'Remark']

DEPRECIATION_METHOD_OPTIONS = ['SLM', 'WDV']
DEPRECIATION_RATE_OPTIONS = [5, 9.5, 10, 13.91, 15, 18.1, 25, 31.67, 33.33, 40, 45, 60, 100]
CAP_AT_SCRAP_OPTIONS = ['Yes', 'No']
ASSET_LIFE_OPTIONS = [1, 2, 3, 4, 5, 6, 8, 10, 15]

ASSET_CLASS_RATE_MAP = {
    'Computers and Servers': [40],
    'Office Equipment': [15],
    'Furniture and Fixtures': [10],
    'Plant and Machinery': [15],
    'Electrical Installations': [10, 15],
    'Vehicles': [15],
    'Mobile Phones': [15],
    'Leasehold Improvements': [10],
    'Buildings': [5, 10],
    'Software / Intangible Assets': [25, 40],
    'Other Tangible Assets': DEPRECIATION_RATE_OPTIONS,
}


@dataclass
class ProcessedAsset:
    asset_id: str
    description: str
    asset_class: str
    location: str
    purchase_date: pd.Timestamp
    depreciation_method: str
    depreciation_rate: float
    assets_life: float
    opening_gross: float
    addition_during_year: float
    deletion_during_year: float
    closing_gross: float
    opening_accum_dep: float
    monthly_dep: Dict[str, float]
    total_dep: float
    dep_on_disposal: float
    closing_accum_dep: float
    opening_net_block: float
    closing_net_block: float
    salvage_value: float
    scrap_floor: float
    status: str
    disposal_date: Optional[pd.Timestamp]
    cap_at_scrap: bool
    sale_value: float = 0.0
    carrying_value_on_sale: float = 0.0
    profit_loss_on_sale: float = 0.0
    vendor: str = ''
    invoice_number: str = ''
    remark: str = ''
    addition_date: Optional[pd.Timestamp] = None


# ---------- Utility ----------
def normalize_header(text: object) -> str:
    if text is None:
        return ''
    return ' '.join(str(text).strip().lower().replace('\n', ' ').split())


def coerce_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors='coerce').fillna(0.0)


def coerce_date(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors='coerce')


def month_bounds(month_end: pd.Timestamp) -> Tuple[pd.Timestamp, pd.Timestamp]:
    start = month_end.replace(day=1)
    return start, month_end


def date_diff_inclusive(start: pd.Timestamp, end: pd.Timestamp) -> int:
    if pd.isna(start) or pd.isna(end) or end < start:
        return 0
    return int((end - start).days) + 1


def first_match(cols: List[str], options: List[str]) -> Optional[str]:
    normalized = {normalize_header(c): c for c in cols}
    for opt in options:
        if opt in normalized:
            return normalized[opt]
    return None


def map_columns(df: pd.DataFrame, config: Dict[str, List[str]]) -> Dict[str, str]:
    mapped = {}
    for canonical, aliases in config.items():
        col = first_match(df.columns.tolist(), aliases)
        if col:
            mapped[canonical] = col
    return mapped


# ---------- Load workbook ----------
def read_uploaded_workbook(uploaded_file) -> Tuple[pd.DataFrame, str, pd.DataFrame]:
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names

    summary_sheet = None
    detail_sheet = None
    for sheet in sheet_names:
        n = normalize_header(sheet)
        if 'summary' in n and summary_sheet is None:
            summary_sheet = sheet
        if ('fy' in n or 'new' in n or 'asset' in n) and detail_sheet is None:
            detail_sheet = sheet

    if detail_sheet is None:
        detail_sheet = sheet_names[-1]
    if summary_sheet is None:
        summary_sheet = sheet_names[0]

    raw = pd.read_excel(uploaded_file, sheet_name=detail_sheet, header=2)
    summary_raw = pd.read_excel(uploaded_file, sheet_name=summary_sheet, header=2)
    return raw, detail_sheet, summary_raw


def prepare_opening_assets(raw: pd.DataFrame) -> pd.DataFrame:
    mapped = map_columns(raw, BASE_REQUIRED_COLS)
    missing = [k for k in ['Assets ID', 'Assets Class', 'Purchase Date', 'Gross Block Closing Value',
                           'Salvage Value', 'Depreciation Method', 'Assets Life',
                           'Closing Gross Block', 'Closing  Accummlated Depreciation'] if k not in mapped]
    if missing:
        raise ValueError(f"Could not identify these columns in the uploaded register: {', '.join(missing)}")

    df = pd.DataFrame()
    for canonical, source in mapped.items():
        df[canonical] = raw[source]

    if 'Assets Description' not in df.columns:
        df['Assets Description'] = ''
    if 'Assets Location' not in df.columns:
        df['Assets Location'] = ''
    if 'Depreciation Rate' not in df.columns:
        df['Depreciation Rate'] = 0.0

    df['Assets ID'] = df['Assets ID'].astype(str).str.strip()
    df = df[df['Assets ID'].ne('') & df['Assets ID'].ne('nan')].copy()

    df['Purchase Date'] = coerce_date(df['Purchase Date'])
    df['Closing Gross Block'] = coerce_numeric(df['Closing Gross Block'])
    df['Closing  Accummlated Depreciation'] = coerce_numeric(df['Closing  Accummlated Depreciation'])
    df['Salvage Value'] = coerce_numeric(df['Salvage Value'])
    df['Depreciation Rate'] = coerce_numeric(df['Depreciation Rate'])
    df['Assets Life'] = coerce_numeric(df['Assets Life']).replace(0, np.nan)
    df['Assets Life'] = df['Assets Life'].fillna(5)

    df['Cap At Scrap Value'] = df['Salvage Value'] > 0
    df['Opening Gross FY26_27'] = df['Closing Gross Block']
    df['Opening Accum FY26_27'] = df['Closing  Accummlated Depreciation']
    df['Opening Net FY26_27'] = df['Opening Gross FY26_27'] - df['Opening Accum FY26_27']
    df['Status FY26 Opening'] = np.where(df['Opening Net FY26_27'] <= df['Salvage Value'], 'At Scrap value', 'Active')

    return df[[
        'Assets ID', 'Assets Description', 'Assets Class', 'Assets Location', 'Purchase Date',
        'Depreciation Method', 'Depreciation Rate', 'Assets Life', 'Salvage Value', 'Cap At Scrap Value',
        'Opening Gross FY26_27', 'Opening Accum FY26_27', 'Opening Net FY26_27', 'Status FY26 Opening'
    ]].copy()




def sanitize_named_range(label: object) -> str:
    text = ''.join(ch if str(ch).isalnum() else '_' for ch in str(label or '').strip())
    while '__' in text:
        text = text.replace('__', '_')
    text = text.strip('_')
    if not text:
        text = 'Blank'
    if text[0].isdigit():
        text = f'N_{text}'
    return text


def is_tangible_asset(asset_class: object) -> bool:
    value = str(asset_class or '').strip().lower()
    if value == '':
        return True
    intangible_markers = ['intangible', 'software', 'license', 'licence', 'goodwill', 'patent', 'trademark']
    return not any(marker in value for marker in intangible_markers)


def apply_additions_defaults(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    class_is_tangible = df['Assets Class'].apply(is_tangible_asset)
    gross = coerce_numeric(df['Gross Block Closing Value'])
    existing_salvage = coerce_numeric(df['Salvage Value'])
    blank_or_zero_salvage = df['Salvage Value'].isna() | existing_salvage.eq(0) | df['Salvage Value'].astype(str).str.strip().eq('')
    df.loc[class_is_tangible & blank_or_zero_salvage, 'Salvage Value'] = (gross[class_is_tangible & blank_or_zero_salvage] * 0.05).round(2)

    cap_series = df['Cap At Scrap Value'].astype(str).str.strip().str.lower()
    blank_cap = cap_series.eq('') | cap_series.eq('nan')
    df.loc[class_is_tangible & blank_cap, 'Cap At Scrap Value'] = 'Yes'
    df.loc[~class_is_tangible & blank_cap, 'Cap At Scrap Value'] = 'No'
    return df

def prepare_additions(additions_file) -> pd.DataFrame:
    if additions_file is None:
        return pd.DataFrame(columns=ADDITION_TEMPLATE_COLS)
    df = pd.read_excel(additions_file)
    for col in ADDITION_TEMPLATE_COLS:
        if col not in df.columns:
            df[col] = '' if col not in ['Gross Block Closing Value', 'Salvage Value', 'Depreciation Rate', 'Assets Life'] else 0
    df = df[ADDITION_TEMPLATE_COLS].copy()
    df['Addition Date'] = coerce_date(df['Addition Date'])
    df['Gross Block Closing Value'] = coerce_numeric(df['Gross Block Closing Value'])
    df['Salvage Value'] = coerce_numeric(df['Salvage Value'])
    df['Depreciation Rate'] = coerce_numeric(df['Depreciation Rate'])
    df['Assets Life'] = coerce_numeric(df['Assets Life']).replace(0, np.nan).fillna(5)
    df = apply_additions_defaults(df)
    df['Cap At Scrap Value'] = df['Cap At Scrap Value'].astype(str).str.strip().str.lower().map({'yes': True, 'true': True, '1': True, 'no': False, 'false': False, '0': False}).fillna(df['Salvage Value'] > 0)
    df['Assets ID'] = df['Assets ID'].astype(str).str.strip()
    df = df[df['Assets ID'].ne('') & df['Assets ID'].ne('nan')].copy()
    return df


def prepare_disposals(disposals_file) -> pd.DataFrame:
    if disposals_file is None:
        return pd.DataFrame(columns=DISPOSAL_TEMPLATE_COLS)
    df = pd.read_excel(disposals_file)
    for col in DISPOSAL_TEMPLATE_COLS:
        if col not in df.columns:
            df[col] = '' if col not in ['Disposed Amount', 'Sale Value'] else 0
    df = df[DISPOSAL_TEMPLATE_COLS].copy()
    df['Disposal Date'] = coerce_date(df['Disposal Date'])
    df['Disposed Amount'] = coerce_numeric(df['Disposed Amount'])
    df['Sale Value'] = coerce_numeric(df['Sale Value'])
    df['Assets ID'] = df['Assets ID'].astype(str).str.strip()
    df = df[df['Assets ID'].ne('') & df['Assets ID'].ne('nan')].copy()
    return df


# ---------- Calculation ----------
def compute_monthly_depreciation(start_date: pd.Timestamp,
                                 end_date: pd.Timestamp,
                                 opening_nb: float,
                                 scrap_floor: float,
                                 remaining_life_end: pd.Timestamp,
                                 cap_at_scrap: bool) -> Tuple[Dict[str, float], float, str]:
    monthly = {label: 0.0 for label in MONTH_LABELS}
    book_value = max(float(opening_nb), 0.0)
    floor = float(scrap_floor) if cap_at_scrap else 0.0

    if pd.isna(start_date):
        start_date = FY_START
    calc_start = max(FY_START, start_date)
    useful_life_end = min(FY_END, remaining_life_end if pd.notna(remaining_life_end) else FY_END)
    posting_end = min(FY_END, end_date if pd.notna(end_date) else FY_END, useful_life_end)

    if book_value <= floor or posting_end < calc_start:
        status = 'At Scrap value' if book_value <= floor and cap_at_scrap else 'No Depreciation'
        return monthly, 0.0, status

    remaining_days_for_rate = date_diff_inclusive(calc_start, useful_life_end)
    if remaining_days_for_rate <= 0:
        return monthly, 0.0, 'No Depreciation'

    depreciable_left = max(book_value - floor, 0.0)
    per_day = depreciable_left / remaining_days_for_rate if remaining_days_for_rate else 0.0

    total_dep = 0.0
    current_value = book_value
    for month_end, label in zip(MONTH_ENDS, MONTH_LABELS):
        m_start, m_end = month_bounds(month_end)
        active_start = max(calc_start, m_start)
        active_end = min(posting_end, m_end)
        if active_end >= active_start:
            days = date_diff_inclusive(active_start, active_end)
            dep = per_day * days
            if cap_at_scrap:
                dep = min(dep, max(current_value - floor, 0.0))
            monthly[label] = round(dep, 2)
            total_dep += monthly[label]
            current_value -= monthly[label]

    closing_nb = book_value - total_dep
    status = 'At Scrap value' if cap_at_scrap and closing_nb <= floor + 0.01 else 'Active'
    return monthly, round(total_dep, 2), status


def process_assets(opening_assets: pd.DataFrame, additions: pd.DataFrame, disposals: pd.DataFrame) -> pd.DataFrame:
    disposal_lookup = {}
    if not disposals.empty:
        disposals = disposals.sort_values('Disposal Date')
        for _, row in disposals.iterrows():
            disposal_lookup[str(row['Assets ID']).strip()] = row

    processed: List[ProcessedAsset] = []

    for _, row in opening_assets.iterrows():
        asset_id = str(row['Assets ID']).strip()
        opening_gross = float(row['Opening Gross FY26_27'])
        opening_accum = float(row['Opening Accum FY26_27'])
        opening_nb = round(opening_gross - opening_accum, 2)
        salvage = float(row['Salvage Value'])
        cap = bool(row['Cap At Scrap Value'])
        purchase_date = pd.to_datetime(row['Purchase Date'])
        life_years = float(row['Assets Life']) if not pd.isna(row['Assets Life']) else 5.0
        life_end = purchase_date + pd.DateOffset(years=int(life_years))

        disposal = disposal_lookup.get(asset_id)
        disposal_date = pd.NaT
        deletion_amount = 0.0
        dep_on_disposal = 0.0
        sale_value = 0.0
        carrying_value_on_sale = 0.0
        profit_loss_on_sale = 0.0
        dep_end = FY_END
        if disposal is not None and pd.notna(disposal['Disposal Date']):
            disposal_date = pd.to_datetime(disposal['Disposal Date'])
            deletion_amount = opening_gross
            sale_value = float(disposal.get('Sale Value', 0.0) or 0.0)
            dep_end = disposal_date

        monthly, total_dep, status = compute_monthly_depreciation(
            start_date=FY_START,
            end_date=dep_end,
            opening_nb=opening_nb,
            scrap_floor=salvage,
            remaining_life_end=life_end,
            cap_at_scrap=cap,
        )

        if pd.notna(disposal_date):
            dep_on_disposal = round(total_dep, 2)
            accumulated_dep_on_sale = round(opening_accum + total_dep, 2)
            carrying_value_on_sale = max(round(opening_gross - accumulated_dep_on_sale, 2), 0.0)
            profit_loss_on_sale = round(sale_value - carrying_value_on_sale, 2)
            closing_accum = 0.0
            closing_gross = 0.0
            closing_nb = 0.0
            status = 'Disposed'
        else:
            closing_accum = round(opening_accum + total_dep, 2)
            closing_gross = round(opening_gross, 2)
            closing_nb = round(closing_gross - closing_accum, 2)

        processed.append(ProcessedAsset(
            asset_id=asset_id,
            description=str(row.get('Assets Description', '') or ''),
            asset_class=str(row.get('Assets Class', '') or ''),
            location=str(row.get('Assets Location', '') or ''),
            purchase_date=purchase_date,
            depreciation_method=str(row.get('Depreciation Method', 'SLM') or 'SLM'),
            depreciation_rate=float(row.get('Depreciation Rate', 0.0) or 0.0),
            assets_life=life_years,
            opening_gross=round(opening_gross, 2),
            addition_during_year=0.0,
            deletion_during_year=round(deletion_amount, 2),
            closing_gross=round(closing_gross, 2),
            opening_accum_dep=round(opening_accum, 2),
            monthly_dep=monthly,
            total_dep=round(total_dep, 2),
            dep_on_disposal=round(dep_on_disposal, 2),
            closing_accum_dep=round(closing_accum, 2),
            opening_net_block=round(opening_nb, 2),
            closing_net_block=round(closing_nb, 2),
            salvage_value=round(salvage, 2),
            scrap_floor=round(salvage if cap else 0.0, 2),
            status=status,
            disposal_date=disposal_date if pd.notna(disposal_date) else None,
            cap_at_scrap=cap,
            sale_value=round(sale_value, 2),
            carrying_value_on_sale=round(carrying_value_on_sale, 2),
            profit_loss_on_sale=round(profit_loss_on_sale, 2),
        ))

    if not additions.empty:
        for _, row in additions.iterrows():
            asset_id = str(row['Assets ID']).strip()
            purchase_date = pd.to_datetime(row['Addition Date'])
            if pd.isna(purchase_date):
                continue
            opening_gross = 0.0
            opening_accum = 0.0
            cost = float(row['Gross Block Closing Value'])
            salvage = float(row['Salvage Value'])
            cap = bool(row['Cap At Scrap Value'])
            life_years = float(row['Assets Life']) if not pd.isna(row['Assets Life']) else 5.0
            life_end = purchase_date + pd.DateOffset(years=int(life_years))
            disposal = disposal_lookup.get(asset_id)
            disposal_date = pd.NaT
            dep_end = FY_END
            deletion_amount = 0.0
            dep_on_disposal = 0.0
            sale_value = 0.0
            carrying_value_on_sale = 0.0
            profit_loss_on_sale = 0.0
            if disposal is not None and pd.notna(disposal['Disposal Date']):
                disposal_date = pd.to_datetime(disposal['Disposal Date'])
                deletion_amount = cost
                sale_value = float(disposal.get('Sale Value', 0.0) or 0.0)
                dep_end = disposal_date

            monthly, total_dep, status = compute_monthly_depreciation(
                start_date=purchase_date,
                end_date=dep_end,
                opening_nb=cost,
                scrap_floor=salvage,
                remaining_life_end=life_end,
                cap_at_scrap=cap,
            )
            if pd.notna(disposal_date):
                dep_on_disposal = round(total_dep, 2)
                accumulated_dep_on_sale = round(total_dep, 2)
                carrying_value_on_sale = max(round(cost - accumulated_dep_on_sale, 2), 0.0)
                profit_loss_on_sale = round(sale_value - carrying_value_on_sale, 2)
                closing_accum = 0.0
                closing_gross = 0.0
                closing_nb = 0.0
                status = 'Added & Disposed'
            else:
                closing_accum = round(total_dep, 2)
                closing_gross = round(cost, 2)
                closing_nb = round(cost - closing_accum, 2)

            processed.append(ProcessedAsset(
                asset_id=asset_id,
                description=str(row.get('Assets Description', '') or ''),
                asset_class=str(row.get('Assets Class', '') or ''),
                location=str(row.get('Assets Location', '') or ''),
                purchase_date=purchase_date,
                depreciation_method=str(row.get('Depreciation Method', 'SLM') or 'SLM'),
                depreciation_rate=float(row.get('Depreciation Rate', 0.0) or 0.0),
                assets_life=life_years,
                opening_gross=0.0,
                addition_during_year=round(cost, 2),
                deletion_during_year=round(deletion_amount, 2),
                closing_gross=round(closing_gross, 2),
                opening_accum_dep=0.0,
                monthly_dep=monthly,
                total_dep=round(total_dep, 2),
                dep_on_disposal=round(dep_on_disposal, 2),
                closing_accum_dep=round(closing_accum, 2),
                opening_net_block=0.0,
                closing_net_block=round(closing_nb, 2),
                salvage_value=round(salvage, 2),
                scrap_floor=round(salvage if cap else 0.0, 2),
                status=status,
                disposal_date=disposal_date if pd.notna(disposal_date) else None,
                cap_at_scrap=cap,
                sale_value=round(sale_value, 2),
                carrying_value_on_sale=round(carrying_value_on_sale, 2),
                profit_loss_on_sale=round(profit_loss_on_sale, 2),
                vendor=str(row.get('Vendor', '') or ''),
                invoice_number=str(row.get('Invoice Number', '') or ''),
                remark=str(row.get('Remark', '') or ''),
                addition_date=purchase_date,
            ))

    detail_rows = []
    for item in processed:
        row = {
            'Assets ID': item.asset_id,
            'Assets Description': item.description,
            'Assets Class': item.asset_class,
            'Assets Location': item.location,
            'Purchase/Add Date': item.addition_date if item.addition_date is not None else item.purchase_date,
            'Original Purchase Date': item.purchase_date,
            'Depreciation Method': item.depreciation_method,
            'Depreciation Rate': item.depreciation_rate,
            'Assets Life (Years)': item.assets_life,
            'Opening Gross Block as on 1-Apr-26': item.opening_gross,
            'Addition During FY 2026-27': item.addition_during_year,
            'Deletion During FY 2026-27': item.deletion_during_year,
            'Closing Gross Block as on 31-Mar-27': item.closing_gross,
            'Opening Accumulated Dep as on 1-Apr-26': item.opening_accum_dep,
        }
        row.update(item.monthly_dep)
        row.update({
            'Total Depreciation FY 2026-27': item.total_dep,
            'Depreciation on Disposal': item.dep_on_disposal,
            'Sale Value on Disposal': item.sale_value,
            'Carrying Value on Sale': item.carrying_value_on_sale,
            'Profit / (Loss) on Sale': item.profit_loss_on_sale,
            'Closing Accumulated Dep as on 31-Mar-27': item.closing_accum_dep,
            'Opening Net Block as on 1-Apr-26': item.opening_net_block,
            'Closing Net Block as on 31-Mar-27': item.closing_net_block,
            'Salvage Value': item.salvage_value,
            'Cap At Scrap Value': 'Yes' if item.cap_at_scrap else 'No',
            'Status': item.status,
            'Disposal Date': item.disposal_date,
            'Vendor': item.vendor,
            'Invoice Number': item.invoice_number,
            'Remark': item.remark,
        })
        detail_rows.append(row)

    detail_df = pd.DataFrame(detail_rows)
    detail_df = detail_df.sort_values(['Assets Class', 'Assets ID']).reset_index(drop=True)
    return detail_df


def build_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    sum_cols = [
        'Opening Gross Block as on 1-Apr-26', 'Addition During FY 2026-27', 'Deletion During FY 2026-27',
        'Closing Gross Block as on 31-Mar-27', 'Opening Accumulated Dep as on 1-Apr-26',
        *MONTH_LABELS,
        'Total Depreciation FY 2026-27', 'Depreciation on Disposal', 'Sale Value on Disposal', 'Carrying Value on Sale', 'Profit / (Loss) on Sale',
        'Closing Accumulated Dep as on 31-Mar-27', 'Opening Net Block as on 1-Apr-26',
        'Closing Net Block as on 31-Mar-27', 'Salvage Value'
    ]
    summary = detail_df.groupby('Assets Class', dropna=False)[sum_cols].sum(numeric_only=True).reset_index()
    grand_total = {'Assets Class': 'TOTAL'}
    for c in sum_cols:
        grand_total[c] = summary[c].sum()
    summary = pd.concat([summary, pd.DataFrame([grand_total])], ignore_index=True)
    return summary


def build_scrap_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    scrap_mask = (
        detail_df['Cap At Scrap Value'].astype(str).str.strip().str.lower().eq('yes')
        & detail_df['Closing Net Block as on 31-Mar-27'].round(2).eq(detail_df['Salvage Value'].round(2))
    )
    scrap_df = detail_df.loc[scrap_mask].copy()

    if scrap_df.empty:
        return pd.DataFrame(columns=[
            'Assets ID', 'Assets Description', 'Assets Class', 'Assets Location',
            'Opening Net Block as on 1-Apr-26', 'Addition During FY 2026-27',
            'Total Depreciation FY 2026-27', 'Closing Net Block as on 31-Mar-27',
            'Salvage Value', 'Status', 'Reached Scrap During FY 2026-27'
        ])

    scrap_df['Reached Scrap During FY 2026-27'] = np.where(
        scrap_df['Opening Net Block as on 1-Apr-26'] > scrap_df['Salvage Value'],
        'Yes',
        np.where(
            (scrap_df['Opening Net Block as on 1-Apr-26'] <= scrap_df['Salvage Value'])
            & (scrap_df['Addition During FY 2026-27'] > 0),
            'Yes',
            'No - Already at Scrap at start'
        )
    )

    cols = [
        'Assets ID', 'Assets Description', 'Assets Class', 'Assets Location',
        'Opening Net Block as on 1-Apr-26', 'Addition During FY 2026-27',
        'Total Depreciation FY 2026-27', 'Closing Net Block as on 31-Mar-27',
        'Salvage Value', 'Status', 'Reached Scrap During FY 2026-27'
    ]
    scrap_df = scrap_df[cols].sort_values(['Reached Scrap During FY 2026-27', 'Assets Class', 'Assets ID']).reset_index(drop=True)

    total_row = {col: '' for col in cols}
    total_row['Assets ID'] = 'TOTAL'
    for col in [
        'Opening Net Block as on 1-Apr-26', 'Addition During FY 2026-27',
        'Total Depreciation FY 2026-27', 'Closing Net Block as on 31-Mar-27', 'Salvage Value'
    ]:
        total_row[col] = scrap_df[col].sum()
    total_row['Reached Scrap During FY 2026-27'] = int((scrap_df['Reached Scrap During FY 2026-27'] == 'Yes').sum())

    scrap_df = pd.concat([scrap_df, pd.DataFrame([total_row])], ignore_index=True)
    return scrap_df


def build_disposal_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    disposal_df = detail_df[detail_df['Status'].astype(str).str.contains('Disposed', case=False, na=False)].copy()
    cols = [
        'Assets ID', 'Assets Description', 'Assets Class', 'Assets Location', 'Disposal Date',
        'Deletion During FY 2026-27', 'Depreciation on Disposal',
        'Carrying Value on Sale', 'Sale Value on Disposal', 'Profit / (Loss) on Sale', 'Status'
    ]
    if disposal_df.empty:
        return pd.DataFrame(columns=cols)

    disposal_df = disposal_df[cols].sort_values(['Disposal Date', 'Assets Class', 'Assets ID']).reset_index(drop=True)
    total_row = {col: '' for col in cols}
    total_row['Assets ID'] = 'TOTAL'
    for col in ['Deletion During FY 2026-27', 'Depreciation on Disposal', 'Carrying Value on Sale', 'Sale Value on Disposal', 'Profit / (Loss) on Sale']:
        total_row[col] = disposal_df[col].sum()
    disposal_df = pd.concat([disposal_df, pd.DataFrame([total_row])], ignore_index=True)
    return disposal_df


# ---------- Export ----------
def style_worksheet(ws):
    header_fill = PatternFill('solid', fgColor='1F4E78')
    sub_fill = PatternFill('solid', fgColor='D9EAF7')
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)
    thin = Side(style='thin', color='D9D9D9')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = 'A2'

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            if isinstance(cell.value, (datetime, pd.Timestamp)):
                cell.number_format = 'dd-mmm-yyyy'
            elif isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    for idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col[:200]:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 24)


def dataframe_to_sheet(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        values = []
        for val in row:
            if pd.isna(val):
                values.append(None)
            elif isinstance(val, pd.Timestamp):
                values.append(val.to_pydatetime())
            else:
                values.append(val)
        ws.append(values)
    style_worksheet(ws)


def build_output_workbook(opening_assets: pd.DataFrame, additions: pd.DataFrame, disposals: pd.DataFrame,
                          detail_df: pd.DataFrame, summary_df: pd.DataFrame, scrap_summary_df: pd.DataFrame, disposal_summary_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = 'README'
    notes = [
        ['Asset Register Automation - FY 2026-27'],
        ['1. Opening balances are picked from the uploaded FY 2025-26 asset register.'],
        ['2. Additions are depreciated from the addition date.'],
        ['3. Sold assets are depreciated up to the disposal date and net block becomes zero after sale.'],
        ['4. Where Cap At Scrap Value = Yes, closing net block is not allowed to go below salvage value for assets not sold.'],
        ['5. Profit / (loss) on sale is computed as Sale Value - (Gross Block - Accumulated Depreciation up to sale date).'],
        ['6. Disposal processing is currently designed as full asset sale at asset-ID level.'],
        ['7. Review the detailed register and summary before finalisation.'],
    ]
    for row in notes:
        ws0.append(row)
    ws0['A1'].font = Font(bold=True, size=14)
    ws0.column_dimensions['A'].width = 110

    dataframe_to_sheet(wb.create_sheet('Opening_FY26_27'), opening_assets)
    dataframe_to_sheet(wb.create_sheet('Additions_Input'), additions if not additions.empty else pd.DataFrame(columns=ADDITION_TEMPLATE_COLS))
    dataframe_to_sheet(wb.create_sheet('Disposals_Input'), disposals if not disposals.empty else pd.DataFrame(columns=DISPOSAL_TEMPLATE_COLS))
    dataframe_to_sheet(wb.create_sheet('FY26_27_Register'), detail_df)
    dataframe_to_sheet(wb.create_sheet('FY26_27_Summary'), summary_df)
    dataframe_to_sheet(wb.create_sheet('Assets_At_Scrap_Value'), scrap_summary_df)
    dataframe_to_sheet(wb.create_sheet('Disposal_Summary'), disposal_summary_df)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_additions_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Additions'
    ws.append(ADDITION_TEMPLATE_COLS)

    list_ws = wb.create_sheet('Lists')
    list_ws.sheet_state = 'hidden'
    list_ws['A1'] = 'Depreciation Method'
    for i, v in enumerate(DEPRECIATION_METHOD_OPTIONS, start=2):
        list_ws[f'A{i}'] = v

    list_ws['B1'] = 'Asset Class'
    asset_classes = list(ASSET_CLASS_RATE_MAP.keys())
    for i, v in enumerate(asset_classes, start=2):
        list_ws[f'B{i}'] = v

    list_ws['C1'] = 'Cap At Scrap Value'
    for i, v in enumerate(CAP_AT_SCRAP_OPTIONS, start=2):
        list_ws[f'C{i}'] = v

    list_ws['D1'] = 'Assets Life'
    for i, v in enumerate(ASSET_LIFE_OPTIONS, start=2):
        list_ws[f'D{i}'] = v

    max_rows = 500
    dv_method = DataValidation(type='list', formula1=f'=Lists!$A$2:$A${len(DEPRECIATION_METHOD_OPTIONS)+1}', allow_blank=True)
    dv_class = DataValidation(type='list', formula1=f'=Lists!$B$2:$B${len(asset_classes)+1}', allow_blank=True)
    dv_cap = DataValidation(type='list', formula1=f'=Lists!$C$2:$C${len(CAP_AT_SCRAP_OPTIONS)+1}', allow_blank=True)
    dv_life = DataValidation(type='list', formula1=f'=Lists!$D$2:$D${len(ASSET_LIFE_OPTIONS)+1}', allow_blank=True)
    for dv in [dv_method, dv_class, dv_cap, dv_life]:
        ws.add_data_validation(dv)

    for row in range(2, max_rows + 2):
        dv_class.add(ws[f'D{row}'])
        dv_method.add(ws[f'H{row}'])
        dv_life.add(ws[f'J{row}'])
        dv_cap.add(ws[f'K{row}'])
        ws[f'G{row}'] = f'=IF(OR(D{row}="",F{row}=""),"",IF(OR(ISNUMBER(SEARCH("intangible",D{row})),ISNUMBER(SEARCH("software",D{row}))),0,ROUND(F{row}*5%,2)))'
        ws[f'K{row}'] = f'=IF(D{row}="","",IF(OR(ISNUMBER(SEARCH("intangible",D{row})),ISNUMBER(SEARCH("software",D{row}))),"No","Yes"))'

    style_worksheet(ws)
    ws.freeze_panes = 'A2'
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def build_disposals_template_xlsx(asset_master: Optional[pd.DataFrame] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Disposals'
    ws.append(DISPOSAL_TEMPLATE_COLS)

    if asset_master is not None and not asset_master.empty:
        lookup = asset_master[['Assets ID', 'Opening Gross FY26_27']].copy()
        lookup = lookup.drop_duplicates(subset=['Assets ID']).sort_values('Assets ID')
        list_ws = wb.create_sheet('Asset_Lookup')
        list_ws.sheet_state = 'hidden'
        list_ws['A1'] = 'Assets ID'
        list_ws['B1'] = 'Gross Block'
        for i, row in enumerate(lookup.itertuples(index=False), start=2):
            list_ws[f'A{i}'] = row[0]
            list_ws[f'B{i}'] = row[1]
        max_rows = max(len(lookup) + 1, 2)
        dv_assets = DataValidation(type='list', formula1=f'=Asset_Lookup!$A$2:$A${max_rows}', allow_blank=True)
        ws.add_data_validation(dv_assets)
        for row in range(2, 502):
            dv_assets.add(ws[f'B{row}'])
            ws[f'C{row}'] = f'=IF(B{row}="","",IFERROR(XLOOKUP(B{row},Asset_Lookup!$A$2:$A${max_rows},Asset_Lookup!$B$2:$B${max_rows},""),""))'
    style_worksheet(ws)
    ws.freeze_panes = 'A2'
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def highlight_scrap_assets(row):
    status = str(row.get('Status', '')).strip().lower()
    reached = str(row.get('Reached Scrap During FY 2026-27', '')).strip().lower()
    if status == 'at scrap value' or reached == 'yes':
        return ['background-color: #fde2e2; color: #b00020; font-weight: 600'] * len(row)
    return [''] * len(row)


# ---------- UI ----------
st.set_page_config(page_title='Asset Register Automation FY 2026-27', layout='wide')
st.title('Asset Register Automation App - FY 2026-27')
st.caption('Upload the FY 2025-26 asset register, then optionally add addition and disposal files for FY 2026-27.')

with st.sidebar:
    st.subheader('Download templates')
    st.download_button(
        'Additions template',
        data=build_additions_template_xlsx(),
        file_name='fy26_27_additions_template.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    st.download_button(
        'Disposals template',
        data=build_disposals_template_xlsx(),
        file_name='fy26_27_disposals_template.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    st.info('Additions template includes dropdowns for asset class, depreciation method, cap at scrap value, and asset life. Depreciation rate is a manual-entry column. Tangible assets default to 5% salvage in the template. Disposal template can auto-pull gross block from selected Asset ID.')

base_file = st.file_uploader('1) Upload base asset register (FY 2025-26 workbook)', type=['xlsx'])
additions_file = st.file_uploader('2) Upload additions for FY 2026-27 (optional)', type=['xlsx'])
disposals_file = st.file_uploader('3) Upload disposals for FY 2026-27 (optional)', type=['xlsx'])

if base_file is not None:
    try:
        raw_df, detail_sheet, summary_raw = read_uploaded_workbook(base_file)
        opening_assets = prepare_opening_assets(raw_df)
        dynamic_disposal_template = build_disposals_template_xlsx(opening_assets)
        additions = prepare_additions(additions_file)
        disposals = prepare_disposals(disposals_file)
        detail_df = process_assets(opening_assets, additions, disposals)
        summary_df = build_summary(detail_df)
        scrap_summary_df = build_scrap_summary(detail_df)
        disposal_summary_df = build_disposal_summary(detail_df)
        output_file = build_output_workbook(opening_assets, additions, disposals, detail_df, summary_df, scrap_summary_df, disposal_summary_df)

        scrap_hits_count = int((scrap_summary_df['Reached Scrap During FY 2026-27'] == 'Yes').sum()) if not scrap_summary_df.empty and 'Reached Scrap During FY 2026-27' in scrap_summary_df.columns else 0
        disposed_count = int((detail_df['Status'].astype(str).str.contains('Disposed', case=False, na=False)).sum())
        net_pl_on_sale = float(detail_df['Profit / (Loss) on Sale'].sum()) if 'Profit / (Loss) on Sale' in detail_df.columns else 0.0

        st.download_button(
            'Download disposal template with Asset ID dropdown',
            data=dynamic_disposal_template,
            file_name='fy26_27_disposals_template_with_asset_ids.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
        k1.metric('Opening assets', int(len(opening_assets)))
        k2.metric('Additions loaded', int(len(additions)))
        k3.metric('Disposals loaded', int(len(disposals)))
        k4.metric('Processed assets', int(len(detail_df)))
        k5.metric('Assets hitting scrap in FY', scrap_hits_count)
        k6.metric('Disposed assets', disposed_count)
        k7.metric('Net profit / (loss) on sale', f"{net_pl_on_sale:,.2f}")

        st.subheader('FY 2026-27 Summary')
        st.dataframe(summary_df, use_container_width=True)

        st.subheader('Assets at Scrap value / Hit Scrap During FY 2026-27')
        if scrap_summary_df.empty:
            st.success('No assets are at scrap value as of 31-Mar-27.')
        else:
            st.dataframe(scrap_summary_df.style.apply(highlight_scrap_assets, axis=1), use_container_width=True)

        st.subheader('Disposal Profit / Loss Summary')
        if disposal_summary_df.empty:
            st.info('No disposals loaded for FY 2026-27.')
        else:
            st.dataframe(disposal_summary_df, use_container_width=True)

        st.subheader('FY 2026-27 Detailed Register')
        detailed_view = detail_df.head(300)
        st.dataframe(detailed_view.style.apply(highlight_scrap_assets, axis=1), use_container_width=True, height=500)
        if len(detail_df) > 300:
            st.caption(f'Showing first 300 rows out of {len(detail_df)} rows.')

        st.download_button(
            'Download automated FY 2026-27 register',
            data=output_file,
            file_name='Asset_Register_Automated_FY_2026_27.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as exc:
        st.error(f'Unable to process the workbook: {exc}')
else:
    st.markdown('''
    **What this app does**

    - picks opening balances from the uploaded FY 2025-26 register
    - computes FY 2026-27 month-wise depreciation from Apr-26 to Mar-27
    - supports in-year additions and disposals
    - additions template has dropdowns for asset class, depreciation method, cap at scrap value, and asset life
    - depreciation rate is a manual-entry column
    - disposal template can auto-fill disposed amount from selected Asset ID gross block
    - stops depreciation at scrap value where required
    - computes profit / loss on sale as Sale Value - (Gross Block - Accumulated Depreciation upto sale date)
    - gives a detailed register, scrap summary, disposal summary, and class-wise summary workbook for download
    ''')
